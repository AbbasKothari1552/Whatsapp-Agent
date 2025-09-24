from openpyxl import load_workbook
from typing import Dict

from src.core.logging_config import get_logger
logger = get_logger(__name__)

from langsmith import traceable

@traceable(name="Excel Parser")
def extract_excel_text(input_path: str) -> Dict:
    """Extract text from Excel files"""
    logger.info(f"Extracting excel file...")
    try:
        wb = load_workbook(input_path)
        text = ""
        
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row if cell) + "\n"
        
        logger.debug(f"Extracted text: {text}")
        
        # with open(output_path, 'w', encoding='utf-8') as f:
        #     f.write(text)
            
        return {
            "method": "openpyxl",
            "word_count": len(text.split()),
            "text": text
        }
    except Exception as e:
        raise Exception(f"Excel extraction failed: {str(e)}")